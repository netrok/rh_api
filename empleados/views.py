# empleados/views.py
from __future__ import annotations
from typing import Optional
from io import BytesIO
from datetime import date

from django.db.models import Q
from django.http import FileResponse
from django_filters import rest_framework as filters

from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiExample, extend_schema

from openpyxl import Workbook
from openpyxl.styles import Font, numbers

from rest_framework import serializers, status, viewsets
from rest_framework import filters as drf_filters  # <-- alias para Search/Ordering
from rest_framework.decorators import action
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response

from core.permissions import IsEmpleadoEditorOrReadOnly, IsRHAdmin
from .models import Empleado
from .serializers import EmpleadoSerializer


# -----------------------
# Filtros
# -----------------------
class EmpleadoFilter(filters.FilterSet):
    """
    Filtros combinables para Empleado.
    - q: búsqueda amplia
    - activo: bool
    - departamento/puesto: por ID
    - departamento_nombre/puesto_nombre: icontains
    - genero: 'M' | 'F' | 'O'
    - deleted: True -> solo borrados, False -> solo vivos, None -> sin alterar queryset base
    """

    q = filters.CharFilter(method="filter_q")
    activo = filters.BooleanFilter()
    departamento = filters.NumberFilter(field_name="departamento_id")
    puesto = filters.NumberFilter(field_name="puesto_id")
    departamento_nombre = filters.CharFilter(
        field_name="departamento__nombre", lookup_expr="icontains"
    )
    puesto_nombre = filters.CharFilter(
        field_name="puesto__nombre", lookup_expr="icontains"
    )
    genero = filters.CharFilter(lookup_expr="iexact")
    deleted = filters.BooleanFilter(method="filter_deleted")

    class Meta:
        model = Empleado
        fields = [
            "activo",
            "departamento",
            "puesto",
            "genero",
            "departamento_nombre",
            "puesto_nombre",
            "deleted",
        ]

    def filter_q(self, queryset, name, value):
        return queryset.filter(
            Q(num_empleado__icontains=value)
            | Q(nombres__icontains=value)
            | Q(apellido_paterno__icontains=value)
            | Q(apellido_materno__icontains=value)
            | Q(email__icontains=value)
            | Q(curp__icontains=value)
            | Q(rfc__icontains=value)
            | Q(nss__icontains=value)
            | Q(departamento__nombre__icontains=value)
            | Q(puesto__nombre__icontains=value)
        )

    def filter_deleted(self, queryset, name, value: Optional[bool]):
        if value is True:
            return queryset.filter(deleted_at__isnull=False)
        if value is False:
            return queryset.filter(deleted_at__isnull=True)
        return queryset


# -----------------------
# ViewSet
# -----------------------
@extend_schema(tags=["Empleados"])
class EmpleadoViewSet(viewsets.ModelViewSet):
    """
    CRUD de Empleados con:
    - soft delete / restore
    - history (django-simple-history)
    - exportación a Excel
    - filtros/ordenación/búsqueda
    """

    serializer_class = EmpleadoSerializer
    permission_classes = [IsEmpleadoEditorOrReadOnly]

    # Backends para filtros/búsqueda/orden
    filterset_class = EmpleadoFilter
    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]

    search_fields = [
        "num_empleado",
        "nombres",
        "apellido_paterno",
        "apellido_materno",
        "email",
        "curp",
        "rfc",
        "nss",
        "departamento__nombre",
        "puesto__nombre",
    ]
    ordering_fields = [
        "num_empleado",
        "fecha_ingreso",
        "apellido_paterno",
        "created_at",
    ]

    parser_classes = (JSONParser, FormParser, MultiPartParser)

    def get_queryset(self):
        """
        Por defecto devuelve solo registros vivos (excluye soft delete).
        Si pasas ?include_deleted=1, parte de todos (vivos + borrados).
        Combina con ?deleted=true|false para filtrar explícitamente.
        """
        include_deleted = self.request.query_params.get("include_deleted")
        has_all = getattr(Empleado, "all_objects", None)
        base_manager = has_all if include_deleted else Empleado.objects
        base = base_manager or Empleado.objects

        return (
            base.select_related("departamento", "puesto", "turno", "horario")
            .all()
            .order_by("num_empleado")
        )

    def get_permissions(self):
        # DELETE/acciones especiales solo Admin (o superuser)
        if self.request.method == "DELETE":
            return [IsRHAdmin()]
        return super().get_permissions()

    # ---------- Acciones: soft delete / restore ----------
    @extend_schema(
        summary="Soft delete",
        description="Marca el empleado como eliminado lógicamente (no se borra físicamente).",
        responses={204: OpenApiTypes.NONE},
    )
    @action(detail=True, methods=["post"], url_path="soft-delete", permission_classes=[IsRHAdmin])
    def soft_delete(self, request: Request, pk: str | None = None) -> Response:
        obj = self.get_object()
        obj.delete()  # soft delete (tu modelo debe implementarlo)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @extend_schema(
        summary="Restaurar",
        description="Restaura un empleado previamente eliminado lógicamente.",
        responses={200: EmpleadoSerializer},
        examples=[OpenApiExample("Restaurado", value={"detail": "ok"})],
    )
    @action(detail=True, methods=["post"], url_path="restore", permission_classes=[IsRHAdmin])
    def restore(self, request: Request, pk: str | None = None) -> Response:
        obj = self.get_object()
        # requiere método restore() en tu modelo/manager
        obj.restore()
        return Response(self.get_serializer(obj).data, status=status.HTTP_200_OK)

    # ---------- Historial ----------
    class _HistoryRecordSerializer(serializers.Serializer):
        history_id = serializers.IntegerField()
        history_date = serializers.DateTimeField()
        history_user = serializers.CharField(allow_null=True)
        history_type = serializers.CharField()  # '+', '~', '-'
        num_empleado = serializers.CharField()
        nombres = serializers.CharField()
        apellidos = serializers.CharField()
        departamento_id = serializers.IntegerField(allow_null=True)
        puesto_id = serializers.IntegerField(allow_null=True)
        activo = serializers.BooleanField()
        deleted_at = serializers.DateTimeField(allow_null=True)

    @extend_schema(
        summary="Historial de cambios",
        description="Devuelve el historial auditado del empleado (django-simple-history).",
        responses={200: _HistoryRecordSerializer(many=True)},
        examples=[
            OpenApiExample(
                "Ejemplo",
                value=[{
                    "history_id": 1,
                    "history_date": "2025-08-24T18:00:00Z",
                    "history_user": "admin",
                    "history_type": "+",
                    "num_empleado": "E001",
                    "nombres": "Juan",
                    "apellidos": "Pérez López",
                    "departamento_id": 1,
                    "puesto_id": 2,
                    "activo": True,
                    "deleted_at": None,
                }],
            )
        ],
    )
    @action(detail=True, methods=["get"], url_path="history")
    def history(self, request: Request, pk: str | None = None) -> Response:
        obj = self.get_object()
        records: list[dict] = []
        # requiere django-simple-history configurado
        for h in obj.history.select_related("history_user").order_by("-history_date"):
            records.append({
                "history_id": h.pk,
                "history_date": h.history_date.isoformat(),
                "history_user": str(h.history_user) if h.history_user else None,
                "history_type": h.history_type,
                "num_empleado": h.num_empleado,
                "nombres": h.nombres,
                "apellidos": f"{getattr(h, 'apellido_paterno', '')} {getattr(h, 'apellido_materno', '')}".strip(),
                "departamento_id": getattr(h, "departamento_id", None),
                "puesto_id": getattr(h, "puesto_id", None),
                "activo": getattr(h, "activo", None),
                "deleted_at": h.deleted_at.isoformat() if getattr(h, "deleted_at", None) else None,
            })
        return Response(records)

    # ---------- Helpers export ----------
    def _apply_front_filters(self, qs):
        """Aplica filtros del front: q, departamento_id, puesto_id, activo."""
        request = self.request
        q = request.query_params.get("q")
        depto_id = request.query_params.get("departamento_id")
        puesto_id = request.query_params.get("puesto_id")
        activo = request.query_params.get("activo")

        if q:
            qs = qs.filter(
                Q(nombres__icontains=q)
                | Q(apellido_paterno__icontains=q)
                | Q(apellido_materno__icontains=q)
                | Q(num_empleado__icontains=q)
                | Q(email__icontains=q)
                | Q(telefono__icontains=q)
            )

        if depto_id:
            qs = qs.filter(departamento_id=depto_id)

        if puesto_id:
            qs = qs.filter(puesto_id=puesto_id)

        if activo is not None:
            val = str(activo).lower()
            if val in {"true", "1", "t", "yes", "y"}:
                qs = qs.filter(activo=True)
            elif val in {"false", "0", "f", "no", "n"}:
                qs = qs.filter(activo=False)

        return qs

    def _base_queryset_for_export(self):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)  # respeta DjangoFilter/Search/Ordering
        qs = self._apply_front_filters(qs)
        return qs.select_related("departamento", "puesto")

    # ---------- Exportación SOLO Excel ----------
    @extend_schema(
        summary="Exportación a Excel",
        description="Descarga un XLSX con el resultado filtrado/ordenado actual.",
        responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY},
    )
    @action(detail=False, methods=["get"], url_path="export/excel")
    def export_excel(self, request: Request, *args, **kwargs) -> FileResponse:
        qs = self._base_queryset_for_export().order_by("id")

        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"

        headers = [
            "ID",
            "Num. empleado",
            "Nombres",
            "Apellido paterno",
            "Apellido materno",
            "Departamento",
            "Puesto",
            "Fecha ingreso",
            "Email",
            "Teléfono",
            "Activo",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        for e in qs:
            ws.append([
                e.id,
                e.num_empleado or "",
                e.nombres or "",
                e.apellido_paterno or "",
                e.apellido_materno or "",
                getattr(getattr(e, "departamento", None), "nombre", "") or "",
                getattr(getattr(e, "puesto", None), "nombre", "") or "",
                e.fecha_ingreso if getattr(e, "fecha_ingreso", None) else "",
                e.email or "",
                getattr(e, "telefono", "") or "",
                "Sí" if getattr(e, "activo", False) else "No",
            ])

        # Filtros y formato de fecha (columna H = 8)
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            row[0].number_format = numbers.FORMAT_DATE_YYYYMMDD2

        # Auto-ancho simple
        for col in ws.columns:
            max_len = 10
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"empleados_{date.today().isoformat()}.xlsx"
        resp = FileResponse(
            bio,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Cache-Control"] = "no-transform"
        return resp
